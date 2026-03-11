#!/usr/bin/env python3
"""Convert Excel (.xlsx) files to Markdown tables with merged cell handling."""

try:
    import openpyxl
except ImportError:
    import subprocess, sys

    try:
        subprocess.check_call(
            [
                sys.executable,
                "-m",
                "pip",
                "install",
                "--user",
                "--break-system-packages",
                "openpyxl",
            ],
            stdout=subprocess.DEVNULL,
        )
    except subprocess.CalledProcessError:
        print(
            "Error: Failed to install openpyxl. Run manually: python3 -m pip install --user openpyxl",
            file=sys.stderr,
        )
        sys.exit(1)
    import openpyxl

import argparse
import sys
from datetime import datetime, date, time
from pathlib import Path


def parse_args():
    p = argparse.ArgumentParser(
        description="Convert Excel (.xlsx) to Markdown tables"
    )
    p.add_argument("file", help="Path to .xlsx file")
    p.add_argument(
        "--sheet",
        action="append",
        dest="sheets",
        help="Sheet name to convert (repeatable). Default: all sheets",
    )
    p.add_argument(
        "--max-rows",
        type=int,
        default=500,
        help="Max data rows per sheet (default: 500)",
    )
    p.add_argument(
        "--no-limit", action="store_true", help="Disable row limit"
    )
    p.add_argument("--output", "-o", help="Output file (default: stdout)")
    p.add_argument(
        "--no-header",
        action="store_true",
        help="Treat first row as data (auto-generate headers)",
    )
    p.add_argument(
        "--header-rows",
        type=int,
        default=1,
        help="Number of header rows for multi-level headers (default: 1)",
    )
    p.add_argument(
        "--header-sep",
        default=" > ",
        help='Multi-level header separator (default: " > ")',
    )
    return p.parse_args()


def build_merge_map(ws):
    """Build a map of (row, col) -> value for all merged cell ranges."""
    merge_map = {}
    for merge_range in ws.merged_cells.ranges:
        top_left = ws.cell(merge_range.min_row, merge_range.min_col).value
        for row in range(merge_range.min_row, merge_range.max_row + 1):
            for col in range(merge_range.min_col, merge_range.max_col + 1):
                merge_map[(row, col)] = top_left
    return merge_map


def get_cell_value(ws, row, col, merge_map):
    """Get cell value, resolving merged cells."""
    if (row, col) in merge_map:
        return merge_map[(row, col)]
    return ws.cell(row, col).value


def format_value(val):
    """Format a cell value for Markdown output."""
    if val is None:
        return ""
    if isinstance(val, datetime):
        return val.strftime("%Y-%m-%d %H:%M:%S")
    if isinstance(val, date):
        return val.strftime("%Y-%m-%d")
    if isinstance(val, time):
        return val.strftime("%H:%M:%S")
    s = str(val)
    s = s.replace("|", "\\|")
    s = s.replace("\n", " ").replace("\r", " ")
    return s.strip()


def build_multi_header(ws, merge_map, num_cols, header_rows, header_sep):
    """Build headers from multiple rows, collapsing duplicates."""
    headers = []
    for col in range(1, num_cols + 1):
        parts = []
        for row in range(1, header_rows + 1):
            val = format_value(get_cell_value(ws, row, col, merge_map))
            parts.append(val)
        # Collapse consecutive duplicates: "인증 > 인증" -> "인증"
        collapsed = []
        for part in parts:
            if not collapsed or part != collapsed[-1]:
                collapsed.append(part)
        headers.append(header_sep.join(p for p in collapsed if p))
    return headers


def sheet_to_markdown(ws, merge_map, args):
    """Convert a single worksheet to Markdown table string."""
    max_col = ws.max_column
    max_row = ws.max_row

    if max_col is None or max_row is None or max_row == 0 or max_col == 0:
        return None, f"Sheet '{ws.title}' is empty"

    if max_col > 20:
        print(
            f"Warning: Sheet '{ws.title}' has {max_col} columns — table may be wide",
            file=sys.stderr,
        )

    # Build headers
    if args.no_header:
        headers = [f"Col{i}" for i in range(1, max_col + 1)]
        data_start = 1
    elif args.header_rows == 1:
        headers = [
            format_value(get_cell_value(ws, 1, col, merge_map))
            for col in range(1, max_col + 1)
        ]
        data_start = 2
    else:
        data_start = args.header_rows + 1
        headers = build_multi_header(
            ws, merge_map, max_col, args.header_rows, args.header_sep
        )

    # Ensure non-empty headers
    headers = [h if h else f"Col{i+1}" for i, h in enumerate(headers)]

    # Build rows
    max_data_rows = max(0, max_row - data_start + 1)
    if args.no_limit:
        limit = max_data_rows
    else:
        limit = min(args.max_rows, max_data_rows)

    rows = []
    for r in range(data_start, data_start + limit):
        row = [
            format_value(get_cell_value(ws, r, col, merge_map))
            for col in range(1, max_col + 1)
        ]
        rows.append(row)

    truncated = max_data_rows - limit

    # Build Markdown
    lines = []
    lines.append("| " + " | ".join(headers) + " |")
    lines.append("| " + " | ".join("---" for _ in headers) + " |")
    for row in rows:
        lines.append("| " + " | ".join(row) + " |")
    if truncated > 0:
        lines.append(f"\n(... {truncated} more rows truncated)")

    return "\n".join(lines), None


def convert(args):
    filepath = Path(args.file)

    if not filepath.exists():
        print(f"Error: File not found: {filepath}", file=sys.stderr)
        sys.exit(1)

    if filepath.suffix.lower() not in (".xlsx", ".xlsm", ".xltx", ".xltm"):
        print(f"Error: Not a valid .xlsx file: {filepath}", file=sys.stderr)
        sys.exit(1)

    try:
        wb = openpyxl.load_workbook(str(filepath), data_only=True)
    except Exception as e:
        msg = str(e)
        if "password" in msg.lower() or "encrypt" in msg.lower():
            print(
                f"Error: File appears to be encrypted/password-protected: {filepath}",
                file=sys.stderr,
            )
        else:
            print(f"Error: Cannot open file: {msg}", file=sys.stderr)
        sys.exit(1)

    # Filter sheets
    if args.sheets:
        sheet_names = []
        for name in args.sheets:
            if name in wb.sheetnames:
                sheet_names.append(name)
            else:
                print(
                    f"Warning: Sheet '{name}' not found. Available: {', '.join(wb.sheetnames)}",
                    file=sys.stderr,
                )
        if not sheet_names:
            print("Error: No valid sheets specified", file=sys.stderr)
            sys.exit(1)
    else:
        sheet_names = wb.sheetnames

    output_parts = []
    for name in sheet_names:
        ws = wb[name]
        merge_map = build_merge_map(ws)
        md, err = sheet_to_markdown(ws, merge_map, args)
        if err:
            print(err, file=sys.stderr)
            continue
        if len(sheet_names) > 1:
            output_parts.append(f"## Sheet: {name}\n\n{md}")
        else:
            output_parts.append(md)

    if not output_parts:
        print("Error: No data to output (all sheets empty)", file=sys.stderr)
        sys.exit(1)

    result = "\n\n".join(output_parts)

    if args.output:
        Path(args.output).write_text(result, encoding="utf-8")
        print(f"Written to {args.output}", file=sys.stderr)
    else:
        print(result)


if __name__ == "__main__":
    args = parse_args()
    convert(args)
